from openpyxl import Workbook, load_workbook

# Para alterar os estilos de formatação das células
from openpyxl.styles import PatternFill, Font

# Carregar dados para variável
wb = load_workbook('C:\\PROJETOS DEV\\PYTON\\test.xlsx')

# Escolhe active sheet
ws = wb.active

# Deleta primeira coluna, que é somente índice
ws.delete_cols(1)

# Cabeçalho em negrito e fundo azul
# Fill parameters(preenchimento)
my_fill = PatternFill(start_color='5399FF',
                   end_color='5399FF',
                   fill_type='solid')
# Bold Parameter
my_font = Font(bold=True)

# Formata o cabeçalho
my_header = ['A1', 'B1', 'C1', 'D1', 'E1',
             'F1', 'G1', 'H1', 'I1', 'J1', 'K1']  # Colunas

for cell in my_header:
    ws[cell].fill = my_fill
    ws[cell].font = my_font

# Adiciona fórmula SUM
ws['K1'] = 'Total' #Celula do resultado
for i in range(2,25): # Variavel de iteração de linhas
    ws['F' + str(i)] = f'=SUM(C{i}:J{i})' #Formula colunas
    ws['F' + str(i)].font = my_font
    ws['F' + str(i)].fill = my_fill

# Salva o arquivo
wb.save('C:\\PROJETOS DEV\\PYTON\\test.xlsx')